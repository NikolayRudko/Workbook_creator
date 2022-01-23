import os
import re
from datetime import datetime

from openpyxl import Workbook


class WorkbookCreator:
    def __init__(self, data_dir):
        self.data_dir = data_dir
        self.files_list = []
        self.files_data = {}

    def run(self) -> None:
        """
        Create Exel file with data of *.txt files
        """
        self.get_files_list()
        self.parce_source_data()
        if self.files_data:
            self.create_workbook()
        else:
            print(f"Folder {self.data_dir} is empty!!!")

    def get_files_list(self) -> None:
        """
        File get list of all files in data_dir folder
        """
        try:
            self.files_list.extend(os.listdir(self.data_dir))
        except FileNotFoundError:
            print(f"Folder {self.data_dir} is not exist!!!")

    def parce_source_data(self) -> None:
        """
        Read each file in file_list and extract data from them to files_data.
        """
        for file in self.files_list:
            if not file.endswith(".txt"):
                continue
            name = self.get_file_name(file)
            source_data = self.read_file(file)
            processed_data = self.get_data(source_data)
            self.files_data[name] = processed_data

    def get_file_name(self, file_name: str) -> str:
        """
        Extracts the file name.

        :param file_name: full name *.txt file
        :return: name without file extension
        """
        pattern = r"(^.*)(\.txt$)"
        result = re.match(pattern, file_name, flags=re.DOTALL)
        return result.group(1)

    def read_file(self, file: str) -> str:
        """
        Read file and return data of file.

        :param file: name file in folder data_dir
        :return: data of file
        """
        file_name = f"{self.data_dir}{os.sep}{file}"
        with open(file_name, "r", encoding="utf-8") as f:
            data = f.read()
        return data

    def get_data(self, source_data: str) -> list:
        """
        Give string and return list expect data.

        :param source_data: source data in file
        :return: list of data without mask
        """
        pattern = r"((.*\])*\s)*((\d{4}\s)*)(mStop.*$)*"
        result = re.match(pattern, source_data, flags=re.DOTALL)
        main_data = result.group(3)
        data = list(map(int, main_data.split()))
        return data

    def create_workbook(self) -> None:
        """
        Create *.xls file, where each sheet have name same *.txt file and consist data of it
        """
        work_book = Workbook()
        default_sheet = work_book.active
        # main_sheet = work_book.create_sheet("Общий", 0)
        for name, value in self.files_data.items():
            ws = work_book.create_sheet(name, -1)
            for i, v in enumerate(value):
                ws.cell(row=i + 1, column=1).value = v
        del work_book['Sheet']
        work_book_name = self.get_name_workbook()
        work_book.save(work_book_name)

    def get_name_workbook(self) -> str:
        """
        Create name for *.xls workbook.
        Name consist of "Workbook" + "current date" + "number of example workbook" + ".xls"
        examples:
        Workbook_22-01-22.xls or Workbook_22-01-22(1).xls

        :return name for workbook
        """
        file_extension = ".xls"
        name = f"Workbook_{self.get_current_date()}"
        files_current_dir = os.listdir(os.getcwd())
        full_name = name + file_extension
        ver = 1
        if full_name not in files_current_dir:
            return full_name
        else:
            while full_name in files_current_dir:
                full_name = f"{name}({ver}){file_extension}"
                ver += 1
        return full_name

    def get_current_date(self) -> str:
        """
        get current day as string in format day-month-year
        example:
        23-01-22

        :return current date
        """
        dt4 = datetime.today()
        date_string1 = dt4.strftime("%d-%m-%y")
        return date_string1


def main():
    dir_name = 'data'
    creator = WorkbookCreator(dir_name)
    creator.run()


if __name__ == "__main__":
    main()
